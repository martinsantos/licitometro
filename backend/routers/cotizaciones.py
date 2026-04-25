"""Cotizaciones CRUD — MongoDB-backed persistence for CotizAR bids."""

import logging
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel

from models.cotizacion import CotizacionCreate
from db.models import cotizacion_entity

logger = logging.getLogger("cotizaciones")

router = APIRouter(
    prefix="/api/cotizaciones",
    tags=["cotizaciones"],
)


from db import get_db
from config.company import DEFAULT_COMPANY_ID


@router.put("/{licitacion_id}")
async def upsert_cotizacion(licitacion_id: str, body: CotizacionCreate, request: Request):
    """Upsert a cotizacion keyed by licitacion_id."""
    db = get_db(request)
    now = datetime.now(timezone.utc)
    data = body.model_dump()
    data["licitacion_id"] = licitacion_id
    data["updated_at"] = now

    existing = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    if existing:
        await db.cotizaciones.update_one(
            {"licitacion_id": licitacion_id},
            {"$set": data},
        )
        updated = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
        return cotizacion_entity(updated)
    else:
        data["created_at"] = now
        result = await db.cotizaciones.insert_one(data)
        doc = await db.cotizaciones.find_one({"_id": result.inserted_id})
        return cotizacion_entity(doc)


@router.get("/")
async def list_cotizaciones(request: Request, enrich: bool = Query(False)):
    """List all cotizaciones, newest first. When enrich=true, add licitacion data."""
    db = get_db(request)
    cursor = db.cotizaciones.find().sort("updated_at", -1).limit(100)
    docs = await cursor.to_list(100)
    results = [cotizacion_entity(d) for d in docs]

    if enrich:
        for item in results:
            lic_id = item.get("licitacion_id")
            if lic_id:
                from db.models import str_to_mongo_id
                lic = await db.licitaciones.find_one(
                    {"_id": str_to_mongo_id(lic_id)},
                    {"opening_date": 1, "budget": 1, "estado": 1},
                )
                if lic:
                    item["opening_date"] = lic.get("opening_date")
                    item["budget"] = lic.get("budget")
                    item["estado"] = lic.get("estado")

    return results


@router.get("/{licitacion_id}")
async def get_cotizacion(licitacion_id: str, request: Request):
    """Get a single cotizacion by licitacion_id."""
    db = get_db(request)
    doc = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    if not doc:
        raise HTTPException(404, "Cotización no encontrada")
    return cotizacion_entity(doc)


@router.delete("/{licitacion_id}")
async def delete_cotizacion(licitacion_id: str, request: Request):
    """Delete a cotizacion."""
    db = get_db(request)
    result = await db.cotizaciones.delete_one({"licitacion_id": licitacion_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Cotización no encontrada")
    return {"deleted": True}


VALID_STATUSES = {"borrador", "presentada", "adjudicada", "rechazada", "perdida", "cancelada"}


class StatusUpdate(BaseModel):
    status: str
    notas_resultado: Optional[str] = None


@router.patch("/{licitacion_id}/status")
async def update_status(licitacion_id: str, body: StatusUpdate, request: Request):
    """Update cotizacion status and optionally record outcome notes."""
    if body.status not in VALID_STATUSES:
        raise HTTPException(400, f"Estado inválido. Válidos: {sorted(VALID_STATUSES)}")
    db = get_db(request)
    now = datetime.now(timezone.utc)
    update: Dict[str, Any] = {"status": body.status, "updated_at": now}
    if body.notas_resultado:
        update["notas_resultado"] = body.notas_resultado
    result = await db.cotizaciones.update_one(
        {"licitacion_id": licitacion_id},
        {"$set": update},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Cotización no encontrada")
    doc = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    return cotizacion_entity(doc)


@router.get("/stats/resumen")
async def stats_resumen(request: Request):
    """Aggregate stats across all cotizaciones: counts + montos by status."""
    db = get_db(request)
    pipeline = [
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "monto_total": {"$sum": "$total"},
        }},
        {"$sort": {"_id": 1}},
    ]
    docs = await db.cotizaciones.aggregate(pipeline).to_list(length=20)
    by_status = {d["_id"]: {"count": d["count"], "monto_total": d["monto_total"]} for d in docs}
    total_count = sum(d["count"] for d in docs)
    total_monto = sum(d["monto_total"] for d in docs)
    adjudicadas = by_status.get("adjudicada", {})
    return {
        "by_status": by_status,
        "total_count": total_count,
        "total_monto": total_monto,
        "adjudicadas_count": adjudicadas.get("count", 0),
        "adjudicadas_monto": adjudicadas.get("monto_total", 0),
        "tasa_exito_pct": round(adjudicadas.get("count", 0) / total_count * 100, 1) if total_count else 0,
    }


class VincularAntecedenteBody(BaseModel):
    antecedente_id: str


@router.post("/{licitacion_id}/vincular-antecedente")
async def vincular_antecedente(licitacion_id: str, body: VincularAntecedenteBody, request: Request):
    """Add an antecedente to a cotizacion."""
    db = get_db(request)
    result = await db.cotizaciones.update_one(
        {"licitacion_id": licitacion_id},
        {"$addToSet": {"antecedentes_vinculados": body.antecedente_id}},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Cotización no encontrada")
    doc = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    return cotizacion_entity(doc)


@router.get("/{licitacion_id}/pdf")
async def generate_pdf(licitacion_id: str, request: Request):
    """Generate professional offer PDF for a cotizacion."""
    db = get_db(request)
    cot = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    if not cot:
        raise HTTPException(404, "Cotizacion not found")

    from db.models import str_to_mongo_id
    try:
        lic = await db.licitaciones.find_one({"_id": str_to_mongo_id(licitacion_id)})
    except Exception:
        lic = None
    if not lic:
        lic = {"title": cot.get("licitacion_title", ""), "organization": cot.get("organization", "")}
    else:
        lic["id"] = str(lic.pop("_id"))

    # Load company profile for brand identity
    company_profile = await db.company_profiles.find_one({"company_id": DEFAULT_COMPANY_ID})

    from services.offer_pdf_chromium import generate_offer_pdf_chromium
    pdf_bytes = generate_offer_pdf_chromium(cot, lic, company_profile)

    filename = f"Oferta_{cot.get('licitacion_title', 'cotizacion')[:40]}.pdf".replace(" ", "_")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/{licitacion_id}/xlsx")
async def export_xlsx(licitacion_id: str, request: Request):
    """Export cotizacion as XLSX with items, subtotals, and company data."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db(request)
    cot = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    if not cot:
        raise HTTPException(404, "Cotización no encontrada")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Header styles
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")

    # Title block
    title = cot.get("licitacion_title", "Cotización")
    org = cot.get("organization", "")
    company = (cot.get("company_data") or {}).get("nombre", "")
    cuit = (cot.get("company_data") or {}).get("cuit", "")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"COTIZACIÓN — {title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Organismo: {org}"
    ws["A3"] = f"Empresa: {company}  |  CUIT: {cuit}"
    ws["A3"].font = Font(italic=True, size=10, color="555555")

    ws.row_dimensions[5].height = 20
    # Column headers
    headers = ["#", "Descripción", "Cantidad", "Unidad", "P.Unitario (ARS)", "Subtotal (ARS)"]
    col_widths = [5, 50, 12, 10, 20, 20]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    items = cot.get("items", [])
    row = 6
    for i, item in enumerate(items, 1):
        desc = item.get("descripcion", "")
        qty = item.get("cantidad", 0)
        unit = item.get("unidad", "u.")
        price = item.get("precio_unitario", 0)
        subtotal = (qty or 0) * (price or 0)

        data = [i, desc, qty, unit, price, subtotal]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            if col_idx in (3, 5, 6):
                cell.alignment = right_align
            if col_idx in (5, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
        row += 1

    # Subtotals
    iva_rate = cot.get("iva_rate", 21)
    subtotal_val = cot.get("subtotal", 0)
    iva_amount = cot.get("iva_amount", 0)
    total_val = cot.get("total", 0)

    row += 1
    for label, value in [("Subtotal sin IVA", subtotal_val), (f"IVA ({iva_rate}%)", iva_amount), ("TOTAL", total_val)]:
        ws.cell(row=row, column=5, value=label).font = bold
        cell = ws.cell(row=row, column=6, value=value)
        cell.font = Font(bold=(label == "TOTAL"), size=12 if label == "TOTAL" else 11)
        cell.number_format = '#,##0.00'
        cell.alignment = right_align
        row += 1

    # Tech data
    tech = cot.get("tech_data") or {}
    if any(tech.values()):
        row += 1
        ws.cell(row=row, column=1, value="Datos Técnicos").font = bold
        row += 1
        for k, v in [("Metodología", tech.get("methodology", "")), ("Plazo", tech.get("plazo", "")), ("Lugar", tech.get("lugar", "")), ("Validez oferta", tech.get("validez", ""))]:
            if v:
                ws.cell(row=row, column=1, value=k).font = Font(italic=True)
                ws.cell(row=row, column=2, value=v)
                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in title[:40])
    filename = f"Cotizacion_{safe_title}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{licitacion_id}/price-intelligence")
async def get_price_intelligence(licitacion_id: str, request: Request):
    """Get price intelligence data for a licitacion."""
    db = get_db(request)
    from services.price_intelligence import get_price_intelligence_service
    service = get_price_intelligence_service(db)
    return await service.get_price_intelligence(licitacion_id)


class Hito(BaseModel):
    id: str = ""
    titulo: str
    fecha: str  # ISO date string YYYY-MM-DD
    completado: bool = False
    notas: Optional[str] = None


class HitosUpdate(BaseModel):
    hitos: List[Hito]


@router.patch("/{licitacion_id}/hitos")
async def update_hitos(licitacion_id: str, body: HitosUpdate, request: Request):
    """Replace the hitos list for a cotizacion."""
    db = get_db(request)
    now = datetime.now(timezone.utc)
    hitos = []
    for hito in body.hitos:
        h = hito.model_dump()
        if not h.get("id"):
            h["id"] = str(uuid.uuid4())[:8]
        hitos.append(h)
    result = await db.cotizaciones.update_one(
        {"licitacion_id": licitacion_id},
        {"$set": {"hitos": hitos, "updated_at": now}},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Cotización no encontrada")
    doc = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    return cotizacion_entity(doc)


@router.delete("/{licitacion_id}/vincular-antecedente/{antecedente_id}")
async def desvincular_antecedente(licitacion_id: str, antecedente_id: str, request: Request):
    """Remove an antecedente from a cotizacion."""
    db = get_db(request)
    result = await db.cotizaciones.update_one(
        {"licitacion_id": licitacion_id},
        {"$pull": {"antecedentes_vinculados": antecedente_id}},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Cotización no encontrada")
    doc = await db.cotizaciones.find_one({"licitacion_id": licitacion_id})
    return cotizacion_entity(doc)
